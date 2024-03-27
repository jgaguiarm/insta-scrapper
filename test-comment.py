import openpyxl
#
# #Criação de Workbook
# workbook = openpyxl.Workbook()
#
# #Mostrar sheets existentes
# print(workbook.sheetnames)
#
# #Para criar novos sheets
# workbook.create_sheet('Ruas')
# workbook.create_sheet('Cidades')
# workbook.create_sheet('Estados')
#
# #Deletando a sheet
# del workbook['Sheet']
#
# #Salvar modificações
# workbook.save('arquivo.xlax')
#
# #Selecionando um Sheet para trabalhar
# sheet_comentarios = workbook['comentarios']
#
# #Sempre colocar cabeçalho
# sheet_comentarios.append([''])


# import re
#
# pattern = re.compile(r'\B@(?!(?:[a-z0-9.]*_){2})(?!(?:[a-z0-9_]*\.){2})[._a-z0-9]{3,24}\b')
# comments = ['@teste','2501','_comment','@01ti']
#
# for comment in comments:
#     if(pattern.match(comment)):
#         print(comment)
#

workbook = openpyxl.Workbook()
del workbook['Sheet']
workbook.create_sheet('Comentarios')
sheet_comentarios = workbook['Comentarios']
sheet_comentarios.append(['Quem comentou','Comentarios'])
workbook.save('comentarios.xlsx')




