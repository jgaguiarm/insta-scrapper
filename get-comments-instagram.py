import tkinter.messagebox

from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from tkinter import messagebox
from selenium.webdriver.common.keys import Keys
import re
import openpyxl

def init_driver():

    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    driver = webdriver.Chrome(options=options)
    driver.get("https://www.instagram.com/")
    driver.implicitly_wait(10)
    return driver

def login(driver):
    loginName = driver.find_element(By.NAME, 'username')
    loginSenha = driver.find_element(By.NAME, 'password')
    loginName.click()
    loginName.send_keys("jaimegaguiar@gmail.com")
    loginSenha.send_keys("Suporte.0413")
    loginSenha.send_keys(Keys.ENTER)
    sleep(5)

def get_comments(driver,workbook):
    driver.get('https://www.instagram.com/p/CzgpCDMiNUp/')
    driver.implicitly_wait(10)

    #Inputando os dados num planilha
    workbook = openpyxl.Workbook()
    del workbook['Sheet']
    workbook.create_sheet('Comentarios')
    sheet_comentarios = workbook['Comentarios']
    sheet_comentarios.append(['Quem comentou','Comentarios'])
    workbook.save('comentarios.xlsx')
    # pattern = re.compile(r'\B@(?!(?:[a-z0-9.]*_){2})(?!(?:[a-z0-9_]*\.){2})[._a-z0-9]{3,24}\b')

    i = 1
    while True:
        comments = driver.find_elements(By.CSS_SELECTOR,
                                     '.x9f619.xjbqb8w.x78zum5.x168nmei.x13lgxp2.x5pf9jr.xo71vjh.xsag5q8.xz9dl7a.x1uhb9sk.x1plvlek.xryxfnj.x1c4vz4f.x2lah0s.x1q0g3np.xqjyukv.x1qjc9v5.x1oa3qoh.x1nhvcw1')

        comments.pop()
        j = 1
        for element in comments:
            j += 1
            if i > j: continue
            print("#" * 40)
            print(f"Comentário {i}")
            i += 1
            elements = element.text.split('\n')
            who = elements[0]
            comment = elements[2]
            print(f'Quem comentou?: {who}')
            print((f'Comentário: {comment}'))
            sheet_comentarios.append([who, comment])
            workbook.save('comentarios.xlsx')
            sleep(1)
            # print(f'Quem comentou?: {elements[0]}')
            # print((f'Comentário: {elements[2]}'))
            #print(element.text)


        if not load_more_comments(driver):
            break


def load_more_comments(driver):
    more_comments = driver.find_element(By.CSS_SELECTOR, 'div[data-visualcompletion="loading-state"]')
    if more_comments:
        try:
            more_comments.click()
            driver.implicitly_wait(10)
            return True
        except:
            return False
    else:
        return False

# driver = init_driver()
# workbook = openpyxl.Workbook()
# login(driver)
# get_comments(driver,workbook)

try:
    driver = init_driver()
    workbook = openpyxl.Workbook()
    login(driver)
    get_comments(driver,workbook)
except:
    #mensagem = "Certifique-se de estar logado no Instagram ao executar o programa!"
    mensagem = "O processamento da aplicação foi interrompido!"
    tkinter.messagebox.showerror("ERRO!",mensagem)
